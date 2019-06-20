
#title           :automation
#description     :automate provisioning control boxes
#author          :Andrew Kubal
#date            :2018
#version         :1
#usage           :python pyscript.py
#notes           :Winnow
#python_version  :3+
#====================================
import subprocess
import os
import time
import sys
import openpyxl
import datetime
from threading import Thread
import gspread
from oauth2client.service_account import ServiceAccountCredentials

def subprocess_cmd(cmd, timeout_sec=1):

    """Execute `cmd` in a subprocess and enforce timeout `timeout_sec` seconds.
    Return subprocess exit code on natural completion of the subprocess.
    Raise an exception if timeout expires before subprocess completes."""
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, shell=False)
    proc_thread = Thread(target=proc.communicate)
    proc_thread.start()
    proc_thread.join(timeout_sec)
    if proc_thread.is_alive():
        # Process still running - kill it and raise timeout error
        try:
            proc.kill()
        except InvalidUserPass as e:
            # The process finished between the `is_alive()` and `kill()`
            return proc.returncode
        # The process was killed
        print('Process #%d ended after %f seconds' % (proc.pid, timeout_sec))
    # Process completed return exit code
    return proc.returncode

def mainexe():
    '''Opening txt file and executing provisioning file. Storing the output.
    Parsing the output. Moving ovpn files to used folder.
    '''

    file_ = open("provision_output.txt", 'w+')
    client = input('\n Enter client name >>> ')

    print("\nStarting provisioning of the system. \n \
    Press ctrl + d to continue\n")

    result = subprocess.run("C:\\Windows\\Sysnative\\bash.exe ;  \
     ls; if [ ! -e provision_tx2.sh ]; then  echo 'File found'; else  echo 'File not found'; fi; \
    ./***redacted*** -d '192.168.90.1' -c %s -u winnowbots -p 'also-heat-among' -v ***redacted*** ; \
    " % (client) , shell=True, stdout=file_)
    file_.close()

    lines = []
    with open ('provision_output.txt', 'rt') as provision_text:
        for line in provision_text:
            lines.append(line)
    #parsing the text.
    i,j,k=0,0,0
    flag=0
    flag2=0
    flag3=0
    dev_id = ''
    vpn_address = ''
    config_data=''
    dev= ['Y','o','u','r','d','e','v','i','c','e','i','d','i','s']
    vpn='Using open vpn certificate '
    config = 'Provisioning'
    for word in lines:
        for letter in word:
            #deviceid scraping
            if letter == dev[i] and i < len(dev)-1:
                i+=1

            if letter == '.' and i > len(dev)-3:
                flag = 1

            if i==len(dev)-1 and flag==0:
                dev_id+=letter

            #vpn scraping
            if letter == vpn[j] and j < len(vpn)-1:
                j+=1

            if letter == 'D' and j > len(vpn)-3:
                flag2 = 1

            if j==len(vpn)-1 and flag2==0:
                vpn_address+=letter

            #ifconfig scraping
            if letter == config[k] and k < len(config)-1:
                k+=1
            else:
                k=0

            if k>=len(config)-2 or flag3 == 1 :
                flag3=1
                config_data+=letter

    print('deviceID:')
    print(dev_id[3:])
    global dev_id_final
    dev_id_final = dev_id[3:]
    print('!!VPN:')
    print('.',vpn_address[2:].replace(' ', '').replace('\n', ''),'.')
    global vpn_address_final
    vpn_address_final = vpn_address[2:]
    vpn_address_final= vpn_address_final.replace(' ', '').replace('\n', '')
    print('config:')
    print(config_data[3:60] + '...')
    global config_data_final
    config_data_final = config_data[2:]
    #move vpn files
    subprocess_cmd("C:\\Windows\\Sysnative\\bash.exe ; \
    cd clients/ ; cd %s/ ; cd pending ; mv %s ../provisioned/ ; \
     exit  " % (client, vpn_address_final) )

def excel():

    ##google-sheets insert
    #create a client to interact with the Google Drive API. A json file is needed.
    scope = ['https://www.googleapis.com/auth/spreadsheets']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('Project auto winnow vision -6ab15661f804.json', scope)
    gc = gspread.authorize(credentials)
    #the url to gsheet
    gsheet = gc.open_by_url('https://docs.google.com/spreadsheets/d/18WX6719-6oZQRa9-tr-yEgMGtxWtIjdhLNo2CVNrPWU/edit#gid=0')
    gworksheet = gsheet.get_worksheet(0)

    #rows 1 to 300 in g-sheet. update when full.
    cell_listB = gworksheet.range('B1:B300')
    print("\nSearching sheet online...\n")

    for gcell in cell_listB:
        if str(gcell.value) is '':
            gcell.value = dev_id_final
            grow=gcell.row
            gcol=gcell.col
            gworksheet.update_cells(cell_listB)
            gworksheet.update_cell(grow,gcol+1,vpn_address_final)
            gworksheet.update_cell(grow,gcol+2,config_data_final)
            time_now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            gcell.value = time_now
            gworksheet.update_cell(grow,gcol+3,time_now)
            break

def end():
    input("\nPress enter or exit window")

if __name__ == "__main__":
    mainexe()
    excel()
    end()

    ##extra lines:
        ##os.startfile( r'C:\Windows\Sysnative\bash.exe' )
        #####--------------------------------
        #subprocess_cmd("C:\\Windows\\Sysnative\\bash.exe ; \
        #cd .. ; cd Downloads/ ; mkdir Provision_folder ;  \
        #cd Provision_folder ; cd wv-provisioning  ; \
        #./provision_tx2.sh -d '192.168.90.1' -c %s -u winnowbots -p 'also-heat-among' -v 'FoodForArtificialThought!' ; \
        #  " % (client) )
        #procedure = subprocess.Popen(command, stdout=subprocess.PIPE, shell=False)
        #processq = procedure.communicate()[0].strip()
        #print(processq)
        #book = openpyxl.load_workbook('Sample.xlsx')


        #book = openpyxl.Workbook()    #to open new excel file
        #print('\nPlease discard the **DeprecationWarning**\n')
        #sheet = book.get_sheet_by_name('Sheet')

        #find empty row
        #r=1
        #while(r<1000):
        #    if sheet.cell(row=r, column=1).value is None:
        #        sheet.cell(row=r, column=1).value = dev_id_final
        #        sheet.cell(row=r, column=2).value = vpn_address_final
        #        sheet.cell(row=r, column=3).value = config_data_final
        #        time_now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        #        sheet.cell(row=r, column=5).value = time_now
        #        break
        #    r+=1
        #book.save('Sample.xlsx')
